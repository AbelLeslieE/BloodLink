"""Administrator donation-history reporting and exports."""

from __future__ import annotations

from datetime import date
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import and_, func, or_, select
from sqlalchemy.orm import Session, joinedload

from backend.auth.dependencies import require_administrator
from backend.database.database import get_db
from backend.database.models import BloodRequest, DonationHistory, Donor, User


router = APIRouter(prefix="/api/donation-history", tags=["donation history"])


def _filtered_donations(
    database_session: Session,
    *,
    search: str | None = None,
    blood_group: str | None = None,
    donation_date: date | None = None,
    district: str | None = None,
    period: str | None = None,
) -> list[DonationHistory]:
    """Return confirmed donations matching the report filters."""
    statement = (
        select(DonationHistory)
        .outerjoin(Donor, DonationHistory.donor_id == Donor.id)
        .join(BloodRequest, DonationHistory.blood_request_id == BloodRequest.id)
        .options(
            joinedload(DonationHistory.donor),
            joinedload(DonationHistory.blood_request),
        )
        .order_by(DonationHistory.donation_date.desc(), DonationHistory.id.desc())
    )

    if search and (term := search.strip()):
        pattern = f"%{term.lower()}%"
        search_fields = [
            Donor.full_name.ilike(pattern),
            Donor.phone.ilike(pattern),
            Donor.email.ilike(pattern),
            Donor.donor_code.ilike(pattern),
            DonationHistory.external_donor_name.ilike(pattern),
        ]
        if term.isdigit():
            search_fields.append(DonationHistory.id == int(term))
        statement = statement.where(or_(*search_fields))

    if blood_group:
        statement = statement.where(
            or_(
                Donor.blood_group == blood_group,
                and_(
                    DonationHistory.external_donor_name.is_not(None),
                    BloodRequest.blood_group == blood_group,
                ),
            )
        )
    if donation_date:
        statement = statement.where(DonationHistory.donation_date == donation_date)
    if district:
        statement = statement.where(Donor.district == district)
    if period == "current_year":
        today = date.today()
        statement = statement.where(
            DonationHistory.donation_date >= date(today.year, 1, 1),
            DonationHistory.donation_date <= date(today.year, 12, 31),
        )

    return list(database_session.scalars(statement).unique())


def _record(donation: DonationHistory) -> dict:
    donor = donation.donor
    is_external = donor is None
    return {
        "id": donation.id,
        "reference": f"DON-{donation.donation_date:%Y}-{donation.id:04d}",
        "donor_name": donor.full_name if donor else donation.external_donor_name or "External donor",
        "donor_code": donor.donor_code if donor else "External",
        "phone": donor.phone if donor else None,
        "blood_group": donor.blood_group if donor else donation.blood_request.blood_group,
        "district": donor.district if donor else "External",
        "source": "external" if is_external else "registered",
        "hospital_name": donation.hospital_name,
        "donation_date": donation.donation_date.isoformat(),
        "points_awarded": donation.points_awarded,
        "status": donation.status,
    }


def _summary(donations: list[DonationHistory]) -> dict:
    total_donations = len(donations)
    total_points = sum(item.points_awarded for item in donations)
    unique_donors = len({
        item.donor_id if item.donor_id is not None else f"external:{item.id}"
        for item in donations
    })
    return {
        "total_donations": total_donations,
        "total_donors": unique_donors,
        "total_points": total_points,
        "average_points": round(total_points / total_donations, 1) if total_donations else 0,
        "confirmed": sum(item.status in {"Donation Confirmed", "Points Awarded"} for item in donations),
        "hospitals": len({item.hospital_name for item in donations if item.hospital_name}),
    }


def _filters(database_session: Session) -> dict:
    blood_groups = database_session.scalars(
        select(func.coalesce(Donor.blood_group, BloodRequest.blood_group))
        .select_from(DonationHistory)
        .outerjoin(Donor, DonationHistory.donor_id == Donor.id)
        .join(BloodRequest, DonationHistory.blood_request_id == BloodRequest.id)
        .distinct()
        .order_by(Donor.blood_group)
    ).all()
    districts = database_session.scalars(
        select(Donor.district)
        .join(DonationHistory, DonationHistory.donor_id == Donor.id)
        .where(Donor.district.is_not(None), Donor.district != "")
        .distinct()
        .order_by(Donor.district)
    ).all()
    return {"blood_groups": blood_groups, "districts": districts}


@router.get("")
def donation_history(
    search: str | None = Query(default=None, max_length=100),
    blood_group: str | None = Query(default=None, max_length=5),
    donation_date: date | None = Query(default=None),
    district: str | None = Query(default=None, max_length=100),
    period: str | None = Query(default=None, pattern="^(current_year)?$"),
    database_session: Session = Depends(get_db),
    _: User = Depends(require_administrator),
) -> dict:
    donations = _filtered_donations(
        database_session,
        search=search,
        blood_group=blood_group,
        donation_date=donation_date,
        district=district,
        period=period,
    )
    records = [_record(donation) for donation in donations]
    return {
        "records": records,
        "summary": _summary(donations),
        "recent_donors": records[:5],
        "filters": _filters(database_session),
    }


def _export_rows(donations: list[DonationHistory]) -> list[list[str | int]]:
    return [
        [
            record["reference"],
            record["donor_name"],
            record["donor_code"],
            record["blood_group"],
            record["district"] or "Not recorded",
            record["hospital_name"],
            record["donation_date"],
            record["points_awarded"],
            record["status"],
        ]
        for donation in donations
        for record in [_record(donation)]
    ]


def _excel_export(rows: list[list[str | int]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Donation History"
    headers = ["Reference", "Donor", "Donor Code", "Blood Group", "District", "Hospital", "Donation Date", "Points", "Status"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="B91C1C")
    for row in rows:
        sheet.append(row)
    for column, width in zip("ABCDEFGHI", [20, 26, 16, 14, 18, 28, 18, 12, 22]):
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _pdf_export(rows: list[list[str | int]]) -> bytes:
    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    table_data = [["Reference", "Donor", "Blood", "District", "Hospital", "Date", "Points", "Status"]]
    table_data.extend([[row[0], row[1], row[3], row[4], row[5], row[6], row[7], row[8]] for row in rows])
    table = Table(table_data, repeatRows=1, colWidths=[24 * mm, 35 * mm, 17 * mm, 25 * mm, 40 * mm, 24 * mm, 16 * mm, 30 * mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#B91C1C")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D1D5DB")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    document.build([Paragraph("BloodLink Donation History", styles["Title"]), Spacer(1, 5 * mm), table])
    return buffer.getvalue()


@router.get("/export/{export_format}")
def export_donation_history(
    export_format: str,
    search: str | None = Query(default=None, max_length=100),
    blood_group: str | None = Query(default=None, max_length=5),
    donation_date: date | None = Query(default=None),
    district: str | None = Query(default=None, max_length=100),
    period: str | None = Query(default=None, pattern="^(current_year)?$"),
    database_session: Session = Depends(get_db),
    _: User = Depends(require_administrator),
) -> Response:
    donations = _filtered_donations(
        database_session,
        search=search,
        blood_group=blood_group,
        donation_date=donation_date,
        district=district,
        period=period,
    )
    rows = _export_rows(donations)
    if export_format == "excel":
        return Response(
            content=_excel_export(rows),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="bloodlink-donation-history.xlsx"'},
        )
    if export_format == "pdf":
        return Response(
            content=_pdf_export(rows),
            media_type="application/pdf",
            headers={"Content-Disposition": 'attachment; filename="bloodlink-donation-history.pdf"'},
        )
    raise HTTPException(status_code=404, detail="Export format must be excel or pdf.")
