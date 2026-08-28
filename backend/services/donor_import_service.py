"""
BloodLink donor Excel import service.
"""

from dataclasses import dataclass
from datetime import date
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from backend.database import crud
from backend.database.schemas import DonorCreate
# ==========================================================
# EXPECTED EXCEL HEADERS
# ==========================================================

HEADER_ALIASES = {
    # BloodLink's original import template.
    "NAME": "full_name",
    "PHONE NUMBER": "phone",
    "CLASS": "class_department",
    "BLOODGROUP": "blood_group",
    "EMAIL-ID": "email",
    "GENDER": "gender",
    "AGE": "age",

    # Columns produced by the BloodLink donor export.
    "FULL NAME": "full_name",
    "PHONE": "phone",
    "BLOOD GROUP": "blood_group",
    "EMAIL": "email",
    "EMAIL ID": "email",
    "DEPARTMENT": "class_department",
    "CLASS / DEPARTMENT": "class_department",
}

REQUIRED_IMPORT_FIELDS = {
    "full_name",
    "phone",
    "blood_group",
}

# ==========================================================
# BLOOD GROUP NORMALIZATION
# ==========================================================

BLOOD_GROUP_MAPPING = {
    "A+": "A+",
    "A+VE": "A+",
    "A +VE": "A+",

    "A-": "A-",
    "A-VE": "A-",
    "A -VE": "A-",

    "B+": "B+",
    "B+VE": "B+",
    "B +VE": "B+",

    "B-": "B-",
    "B-VE": "B-",
    "B -VE": "B-",

    "AB+": "AB+",
    "AB+VE": "AB+",
    "AB +VE": "AB+",

    "AB-": "AB-",
    "AB-VE": "AB-",
    "AB -VE": "AB-",

    "O+": "O+",
    "O+VE": "O+",
    "O +VE": "O+",

    "O-": "O-",
    "O-VE": "O-",
    "O -VE": "O-",
}
ALLOWED_BLOOD_GROUPS = {"A+", "A-", "B+", "B-", "AB+", "AB-", "O+", "O-"}
# ==========================================================
# IMPORT SUMMARY
# ==========================================================

@dataclass
class ImportSummary:
    total_rows: int = 0
    imported: int = 0
    duplicates: int = 0
    skipped: int = 0

    errors: list[dict] | None = None

    def __post_init__(self):

        if self.errors is None:
            self.errors = []


# ==========================================================
# LOAD WORKBOOK
# ==========================================================

def load_excel_workbook(
    file_bytes: bytes,
) -> Worksheet:
    """
    Load the uploaded Excel workbook.
    """

    workbook = load_workbook(
        BytesIO(file_bytes),
        data_only=True,
        read_only=True,
    )

    return workbook.active


# ==========================================================
# READ HEADER ROW
# ==========================================================

def build_header_map(
    worksheet: Worksheet,
) -> tuple[dict[str, int], int]:
    """
    Locate the header row and map supported header names to columns.

    Legacy import files place headers on row two, while files exported
    from BloodLink place headers on row one. Supporting both means an
    exported workbook can be imported again without manual edits.
    """

    for row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=1,
            max_row=min(10, worksheet.max_row),
        ),
        start=1,
    ):

        header_map: dict[str, int] = {}

        for index, cell in enumerate(row, start=1):

            normalized_header = (
                " ".join(str(cell.value).strip().upper().split())
                if cell.value is not None
                else ""
            )

            field_name = HEADER_ALIASES.get(normalized_header)

            if field_name:
                header_map[field_name] = index

        if REQUIRED_IMPORT_FIELDS.issubset(header_map):
            return header_map, row_number

    raise ValueError(
        "Missing required headers: Full Name, Phone, and Blood Group."
    )
# ==========================================================
# CELL HELPERS
# ==========================================================

def get_cell_value(
    row,
    header_map: dict[str, int],
    column_name: str,
):
    """
    Return a trimmed cell value.

    Missing columns return None.
    Blank cells return None.
    """

    column_index = header_map.get(column_name)

    if column_index is None:
        return None

    value = row[column_index - 1].value

    if value is None:
        return None

    value = str(value).strip()

    if value == "":
        return None

    return value


# ==========================================================
# NORMALIZE BLOOD GROUP
# ==========================================================

def normalize_blood_group(
    value: str | None,
) -> str | None:

    if value is None:
        return None

    cleaned = (
        value.upper()
        .replace(" ", "")
    )

    return BLOOD_GROUP_MAPPING.get(
        cleaned,
        cleaned,
    )


# ==========================================================
# EMPTY ROW CHECK
# ==========================================================

def row_is_empty(
    row,
) -> bool:

    for cell in row:

        if cell.value not in (
            None,
            "",
        ):
            return False

    return True
# ==========================================================
# PARSE ONE DONOR ROW
# ==========================================================

def parse_donor_row(
    row,
    header_map: dict[str, int],
) -> dict:

    blood_group = normalize_blood_group(
        get_cell_value(
            row,
            header_map,
            "blood_group",
        )
    )

    age = get_cell_value(
        row,
        header_map,
        "age",
    )

    try:
        parsed_age = int(age) if age else None
    except (TypeError, ValueError):
        parsed_age = None

    return {

        "full_name": get_cell_value(
            row,
            header_map,
            "full_name",
        ),

        "phone": get_cell_value(
            row,
            header_map,
            "phone",
        ),

        "blood_group": blood_group,

        "class_department": get_cell_value(
            row,
            header_map,
            "class_department",
        ),

        "email": get_cell_value(
            row,
            header_map,
            "email",
        ),

        "gender": get_cell_value(
            row,
            header_map,
            "gender",
        ),

        "age": parsed_age,

    }
# ==========================================================
# IMPORT DONORS FROM EXCEL
# ==========================================================

def import_donors_from_excel(
    database_session: Session,
    file_bytes: bytes,
) -> ImportSummary:
    """
    Import donors from an Excel workbook.

    Only valid rows are imported.
    Invalid or duplicate rows are skipped.
    """

    worksheet = load_excel_workbook(file_bytes)

    header_map, header_row = build_header_map(
        worksheet
    )

    summary = ImportSummary()

    # ------------------------------------------------------
    # Data starts immediately after the detected header row.
    # ------------------------------------------------------

    for excel_row_number, row in enumerate(
        worksheet.iter_rows(min_row=header_row + 1),
        start=header_row + 1,
    ):

        # ----------------------------------------------
        # Ignore empty rows
        # ----------------------------------------------

        if row_is_empty(row):
            continue

        summary.total_rows += 1

        donor = parse_donor_row(
            row,
            header_map,
        )

        # ----------------------------------------------
        # Required fields
        # ----------------------------------------------

        if not donor["full_name"]:

            summary.skipped += 1

            summary.errors.append({
                "row": excel_row_number,
                "reason": "Missing donor name",
            })

            continue

        if not donor["phone"]:

            summary.skipped += 1

            summary.errors.append({
                "row": excel_row_number,
                "reason": "Missing phone number",
            })

            continue

        if not donor["blood_group"]:

            summary.skipped += 1

            summary.errors.append({
                "row": excel_row_number,
                "reason": "Missing blood group",
            })

            continue

        if donor["blood_group"] not in ALLOWED_BLOOD_GROUPS:
            summary.skipped += 1
            summary.errors.append({
                "row": excel_row_number,
                "reason": "Invalid blood group",
            })
            continue

        # ----------------------------------------------
        # Duplicate phone
        # ----------------------------------------------

        existing_phone = crud.get_donor_by_phone(
            database_session,
            donor["phone"],
        )

        if existing_phone:

            summary.duplicates += 1

            continue

        # ----------------------------------------------
        # Duplicate email
        # ----------------------------------------------

        if donor["email"]:

            existing_email = crud.get_donor_by_email(
                database_session,
                donor["email"],
            )

            if existing_email:

                summary.duplicates += 1

                continue

        # ----------------------------------------------
        # Create donor schema
        # ----------------------------------------------

        donor_schema = DonorCreate(

            full_name=donor["full_name"],

            blood_group=donor["blood_group"],

            gender=donor["gender"] or "Not Specified",

            age=donor["age"],

            phone=donor["phone"],

            email=donor["email"],

            class_department=donor["class_department"],

            date_of_birth=None,

            weight=None,

            district=None,

            city=None,

            address=None,

            emergency_contact_name=None,

            emergency_contact_phone=None,

            last_donation_date=None,

            total_donations=0,

            status="Available",

            hb_above_12_5="Not Recorded",

            regular_medication="Not Recorded",

            bp_normal="Not Recorded",

            remarks="Imported from Excel",

        )

        # ----------------------------------------------
        # Save donor
        # ----------------------------------------------

        crud.create_donor(
            database_session,
            donor_schema,
        )

        summary.imported += 1

    return summary
